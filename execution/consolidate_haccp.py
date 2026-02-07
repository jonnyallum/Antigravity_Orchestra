"""
Consolidate ALL HACCP data (old + new) into ONE single spreadsheet.
All 5 sections on ONE sheet, new data highlighted green.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill
import csv
import os

SRC = r'c:\Users\jonny\Desktop\Jai.OS 4.0 template\.tmp\haccp_sheets\haccp_logbook.xlsx'
CSV_DIR = r'c:\Users\jonny\Desktop\Jai.OS 4.0 template\.tmp\haccp_sheets'
OUT = r'c:\Users\jonny\Desktop\HACCP_ALL_DATA_CONSOLIDATED.xlsx'

# Styles
title_font = Font(bold=True, size=14, color='FFFFFF')
title_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
section_font = Font(bold=True, size=12, color='FFFFFF')
section_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
new_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

# Load original
orig_wb = openpyxl.load_workbook(SRC, data_only=True)
print(f"Loaded original: {orig_wb.sheetnames}")

# Create new workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'All HACCP Data'

row = 1

# Title
ws.cell(row=row, column=1, value='MANOR BARN CARE HOME — HACCP LOGBOOK (CONSOLIDATED)').font = title_font
ws.cell(row=row, column=1).fill = title_fill
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
row += 1
ws.cell(row=row, column=1, value='Old data + New fill data (green rows) from 18 Dec 2025 to 6 Feb 2026')
row += 2

def write_section(ws, row, title, headers, orig_sheet_name, csv_file, max_cols):
    """Write a section with old + new data."""
    # Section header
    ws.cell(row=row, column=1, value=title).font = section_font
    ws.cell(row=row, column=1).fill = section_fill
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
    row += 1
    
    # Column headers
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
    row += 1
    
    # Old data from original
    old_count = 0
    if orig_sheet_name in orig_wb.sheetnames:
        orig_ws = orig_wb[orig_sheet_name]
        for orig_row in orig_ws.iter_rows(min_row=2, values_only=True):
            vals = list(orig_row)[:max_cols]
            if any(v is not None for v in vals):
                for col, val in enumerate(vals, 1):
                    if val is not None:
                        ws.cell(row=row, column=col, value=val)
                row += 1
                old_count += 1
    
    # New data from CSV (green)
    new_count = 0
    csv_path = os.path.join(CSV_DIR, csv_file)
    if os.path.exists(csv_path):
        with open(csv_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            next(reader)  # skip header
            for csv_row in reader:
                for col, val in enumerate(csv_row, 1):
                    c = ws.cell(row=row, column=col)
                    try:
                        num = float(val)
                        c.value = int(num) if num == int(num) else num
                    except (ValueError, TypeError):
                        c.value = val
                    c.fill = new_fill
                row += 1
                new_count += 1
    
    print(f"  {title}: {old_count} old + {new_count} new = {old_count + new_count} total")
    return row + 2  # gap between sections

# Section 1: Fridge & Freezer
row = write_section(ws, row,
    'SECTION 1: FRIDGE & FREEZER TEMP LOG',
    ['Date', 'Check (AM/PM)', 'Fridge 1', 'Fridge 2', 'Fridge 3', 'Freezer 1', 'Freezer 2', 'Notes'],
    'Fridge & Freezer Temp Log',
    'fridge_freezer_fill.csv', 8)

# Section 2: Cooking Temps
row = write_section(ws, row,
    'SECTION 2: COOKING TEMPS',
    ['Date', 'Time', 'Shift', 'Meal', 'Location/Equipment', 'Target Min (°C)', 'Probe 1 (°C)', 'Probe 2 (°C)', 'Probe 3 (°C)', 'Pass/Fail', 'Corrective Action', 'Verified by'],
    'Cooking Temps',
    'cooking_temps_fill.csv', 12)

# Section 3: Cleaning Schedule
row = write_section(ws, row,
    'SECTION 3: CLEANING SCHEDULE',
    ['Date', 'Area / Equipment Cleaned', 'Cleaning Product Used', 'Staff Member', 'Supervisor Check', 'Comments'],
    'Cleaning Schedule',
    'cleaning_schedule_fill.csv', 6)

# Section 4: Weekly Sign-off
row = write_section(ws, row,
    'SECTION 4: WEEKLY SIGN-OFF',
    ['Week Commencing (Mon)', 'Chef Sign', 'Notes'],
    'Weekly Sign-off',
    'weekly_signoff_fill.csv', 3)

# Section 5: Probe Calibration
row = write_section(ws, row,
    'SECTION 5: PROBE CALIBRATION',
    ['Date', 'Probe ID', 'Ice Slurry (°C)', 'Boiling Water (°C)', 'Within ±0.5°C?', 'Action Taken', 'Initials', 'Notes'],
    'Probe Calibration',
    'probe_calibration_fill.csv', 8)

# Column widths
ws.column_dimensions['A'].width = 18
ws.column_dimensions['B'].width = 16
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 38
ws.column_dimensions['E'].width = 22
ws.column_dimensions['F'].width = 16
ws.column_dimensions['G'].width = 14
ws.column_dimensions['H'].width = 50

wb.save(OUT)
print(f"\nSaved: {OUT}")
print(f"Total rows: {row}")
print(f"File size: {os.path.getsize(OUT):,} bytes")
print("\nDone! All 5 sections on ONE sheet. New data in GREEN.")
